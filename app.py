import os
import io
import csv
import json
import time
import socket
import logging
import secrets as secrets_module
import string

from datetime import datetime, timedelta, timezone
from functools import wraps
from io import BytesIO

from flask import (Flask, render_template, request, jsonify, flash,
                   redirect, url_for, send_file, session, g)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# ── App factory ────────────────────────────────────────────────────────────────

def create_app(env=None):
    app = Flask(__name__)

    # Load config from config.py — no more inline hardcoding
    from config import config
    env = env or os.environ.get('FLASK_ENV', 'default')
    cfg = config.get(env, config['default'])
    app.config.from_object(cfg)
    cfg.init_app(app)

    # Register extensions
    from models import db, User, Book, BookBorrowing, Log
    db.init_app(app)

    mail_instance = Mail(app)

    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'Please log in to access this page.'
    login_manager.login_message_category = 'warning'

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    # Create tables and default admin — retry loop handles slow cloud DB startup
    import time
    with app.app_context():
        for attempt in range(10):
            try:
                db.create_all()
                admin = User.query.filter_by(username='admin').first()
                if not admin:
                    admin = User(
                        username='admin',
                        email='admin@school.edu',
                        full_name='System Administrator',
                        role='admin'
                    )
                    admin.set_password('admin123')
                    db.session.add(admin)
                    db.session.commit()
                    print("✓ Default admin created  |  user: admin  |  pass: admin123")
                break  # success — exit retry loop
            except Exception as e:
                if attempt < 9:
                    print(f"DB not ready (attempt {attempt+1}/10), retrying in 3s… ({e})")
                    time.sleep(3)
                else:
                    print(f"ERROR: Could not connect to database after 10 attempts: {e}")
                    raise

    # Register all routes (passing mail and db references)
    _register_routes(app, db, mail_instance, User, Book, BookBorrowing, Log)

    return app, mail_instance


# ── Helpers ────────────────────────────────────────────────────────────────────

def utcnow():
    return datetime.now(timezone.utc)


def _log(db, Log, user_id, action, details, ip):
    """Create an audit log entry (does NOT commit — caller must commit)."""
    entry = Log(user_id=user_id, action=action, details=details, ip_address=ip)
    db.session.add(entry)


def _book_query_from_args(Book, args):
    """Build a filtered Book query from request args."""
    q = Book.query
    search = args.get('search', '')
    subject = args.get('subject', '')
    grade = args.get('grade', '')
    avail = args.get('available', '')

    if search:
        q = q.filter(
            (Book.title.ilike(f'%{search}%')) |
            (Book.author.ilike(f'%{search}%')) |
            (Book.book_number.ilike(f'%{search}%'))
        )
    if subject:
        q = q.filter_by(subject=subject)
    if grade:
        q = q.filter_by(grade_level=grade)
    if avail == 'available':
        q = q.filter_by(is_available=True)
    elif avail == 'borrowed':
        q = q.filter_by(is_available=False)
    return q, search, subject, grade, avail


def _subjects_and_grades(db, Book):
    subjects = [s[0] for s in db.session.query(Book.subject).distinct().all() if s[0]]
    grades = [g[0] for g in db.session.query(Book.grade_level).distinct().all() if g[0]]
    return subjects, grades


def _parse_book_row(row_data):
    """Extract and validate a book dict from an import row. Returns (data, errors)."""
    required = ['book_number', 'title', 'author', 'subject', 'grade_level']
    missing = [f for f in required if not row_data.get(f)]
    if missing:
        return None, f"Missing fields: {', '.join(missing)}"
    return {k: (str(v).strip() if v is not None else '') for k, v in row_data.items()}, None


def _upsert_book(Book, db, row_data, skip_duplicates, overwrite_existing):
    """Create or update a book from a row dict. Returns ('created'|'updated'|'skipped', msg)."""
    data, err = _parse_book_row(row_data)
    if err:
        return 'skipped', err

    existing = Book.query.filter_by(book_number=data['book_number']).first()
    if existing:
        if not overwrite_existing:
            if skip_duplicates:
                return 'skipped', f"Book '{data['book_number']}' already exists"
        if overwrite_existing:
            for field in ['title', 'author', 'subject', 'grade_level', 'isbn',
                          'publisher', 'edition', 'shelf_location']:
                setattr(existing, field, data.get(field, ''))
            existing.total_copies = Book._parse_int(data.get('total_copies', 1))
            existing.copies_available = Book._parse_int(
                data.get('copies_available', existing.total_copies), existing.total_copies)
            existing.publication_year = Book._parse_year(data.get('publication_year'))
            existing.is_available = existing.copies_available > 0
            return 'updated', None

    book = Book(
        book_number=data['book_number'],
        title=data['title'],
        author=data['author'],
        subject=data['subject'],
        grade_level=data['grade_level'],
        isbn=data.get('isbn', ''),
        publisher=data.get('publisher', ''),
        edition=data.get('edition', ''),
        shelf_location=data.get('shelf_location', ''),
    )
    book.total_copies = Book._parse_int(data.get('total_copies', 1))
    book.copies_available = Book._parse_int(
        data.get('copies_available', book.total_copies), book.total_copies)
    book.publication_year = Book._parse_year(data.get('publication_year'))
    book.is_available = book.copies_available > 0
    db.session.add(book)
    return 'created', None


# ── Route registration ─────────────────────────────────────────────────────────

def _register_routes(app, db, mail, User, Book, BookBorrowing, Log):

    logger = logging.getLogger(__name__)

    # ── Decorators ─────────────────────────────────────────────────────────────

    def admin_required(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated or not current_user.is_admin():
                flash('Admin access required', 'danger')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated

    def librarian_required(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated or \
                    current_user.role not in ['admin', 'librarian']:
                flash('Librarian access required', 'danger')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated

    # ── Context processors ────────────────────────────────────────────────────

    @app.context_processor
    def inject_now():
        # Use a callable so templates get a fresh datetime each render
        return {'datetime': datetime, 'now': datetime.now}

    # ── Before/after request ──────────────────────────────────────────────────

    @app.before_request
    def check_idle_timeout():
        if current_user.is_authenticated:
            last_activity = session.get('last_activity')
            if last_activity:
                last_activity = datetime.fromisoformat(last_activity)
                if datetime.now() - last_activity > app.config['IDLE_TIMEOUT']:
                    logout_user()
                    session.clear()
                    flash('You have been logged out due to inactivity.', 'info')
                    return redirect(url_for('login'))
            session['last_activity'] = datetime.now().isoformat()

    @app.after_request
    def add_header(response):
        if current_user.is_authenticated:
            response.headers['Cache-Control'] = \
                'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '-1'
        return response

    # ── Token helpers ─────────────────────────────────────────────────────────

    def generate_reset_token(email):
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        return s.dumps(email, salt=app.config['SECURITY_PASSWORD_SALT'])

    def verify_reset_token(token, expiration=3600):
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        try:
            return s.loads(token, salt=app.config['SECURITY_PASSWORD_SALT'],
                           max_age=expiration)
        except (SignatureExpired, BadSignature):
            return None

    # ── Email helpers ─────────────────────────────────────────────────────────

    def send_borrow_notification(borrowing, book, borrower, is_bulk=False, books_list=None):
        try:
            pdf_buffer, receipt_number = generate_borrow_receipt_pdf(
                borrowing, book, borrower, is_bulk, books_list)

            if is_bulk and books_list:
                subject = f'Bulk Book Borrowing Confirmation - {len(books_list)} Books (Receipt #{receipt_number})'
                template = 'bulk_borrow_notification_email.html'
                ctx = {'borrower': borrower, 'books': books_list,
                       'borrowing_date': utcnow(), 'due_date': borrowing.due_date,
                       'current_user': current_user, 'receipt_number': receipt_number}
            else:
                subject = f'Book Borrowing Confirmation: {book.title} (Receipt #{receipt_number})'
                template = 'borrow_notification_email.html'
                ctx = {'borrower': borrower, 'book': book, 'borrowing': borrowing,
                       'current_user': current_user, 'receipt_number': receipt_number}

            msg = Message(subject=subject, recipients=[borrower.email])
            msg.attach(f"Borrow_Receipt_{receipt_number}.pdf", "application/pdf",
                       pdf_buffer.getvalue())
            msg.html = render_template(template, **ctx)
            mail.send(msg)

            _log(db, Log, borrower.id, 'EMAIL_NOTIFICATION',
                 f'Borrow notification sent to {borrower.email} (Receipt: {receipt_number})',
                 request.remote_addr)
            db.session.commit()
            return True, receipt_number
        except Exception as e:
            logger.error(f"Error sending borrow email: {e}")
            _log(db, Log, getattr(borrower, 'id', None), 'EMAIL_ERROR',
                 f'Failed to send borrow email: {e}', request.remote_addr)
            db.session.commit()
            return False, None

    def send_return_notification(borrowing, book, borrower, is_bulk=False,
                                 returned_books_list=None, total_fine=0, fines_paid=0):
        try:
            pdf_buffer, receipt_number = generate_return_receipt_pdf(
                borrowing, book, borrower, is_bulk, returned_books_list, total_fine, fines_paid)

            if is_bulk and returned_books_list:
                subject = f'Bulk Book Return Confirmation - {len(returned_books_list)} Books (Receipt #{receipt_number})'
                template = 'bulk_return_notification_email.html'
                ctx = {'borrower': borrower, 'returned_books': returned_books_list,
                       'return_date': utcnow(), 'total_fine': total_fine,
                       'fines_paid': fines_paid, 'current_user': current_user,
                       'receipt_number': receipt_number}
            else:
                subject = f'Book Return Confirmation: {book.title} (Receipt #{receipt_number})'
                template = 'return_notification_email.html'
                ctx = {'borrower': borrower, 'book': book, 'borrowing': borrowing,
                       'fine_amount': borrowing.fine_amount, 'fine_paid': borrowing.fine_paid,
                       'current_user': current_user, 'receipt_number': receipt_number}

            msg = Message(subject=subject, recipients=[borrower.email])
            msg.attach(f"Return_Receipt_{receipt_number}.pdf", "application/pdf",
                       pdf_buffer.getvalue())
            msg.html = render_template(template, **ctx)
            mail.send(msg)

            _log(db, Log, borrower.id, 'EMAIL_NOTIFICATION',
                 f'Return notification sent to {borrower.email} (Receipt: {receipt_number})',
                 request.remote_addr)
            db.session.commit()
            return True, receipt_number
        except Exception as e:
            logger.error(f"Error sending return email: {e}")
            _log(db, Log, getattr(borrower, 'id', None), 'EMAIL_ERROR',
                 f'Failed to send return email: {e}', request.remote_addr)
            db.session.commit()
            return False, None

    # ── Auth routes ───────────────────────────────────────────────────────────

    @app.route('/')
    def index():
        return redirect(url_for('dashboard') if current_user.is_authenticated else url_for('login'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            user = User.query.filter_by(username=username, is_active=True).first()
            if user and user.check_password(password):
                login_user(user)
                logger.info(f"User {username} logged in")
                _log(db, Log, user.id, 'LOGIN', f'User {username} logged in', request.remote_addr)
                db.session.commit()
                flash('Logged in successfully!', 'success')
                return redirect(url_for('dashboard'))
            flash('Invalid username or password', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        _log(db, Log, current_user.id, 'LOGOUT',
             f'User {current_user.username} logged out', request.remote_addr)
        db.session.commit()
        logout_user()
        flash('Logged out successfully!', 'success')
        return redirect(url_for('login'))

    @app.route('/change-password', methods=['GET', 'POST'])
    @login_required
    def change_password():
        if request.method == 'POST':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            if not current_user.check_password(current_password):
                flash('Current password is incorrect', 'danger')
            elif new_password != confirm_password:
                flash('New passwords do not match', 'danger')
            elif len(new_password) < 8:
                flash('Password must be at least 8 characters', 'danger')
            else:
                current_user.set_password(new_password)
                _log(db, Log, current_user.id, 'PASSWORD_CHANGE',
                     'User changed password', request.remote_addr)
                db.session.commit()
                flash('Password changed successfully!', 'success')
                return redirect(url_for('profile'))
        return render_template('change_password.html')

    @app.route('/forgot-password', methods=['GET', 'POST'])
    def forgot_password():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        if request.method == 'POST':
            email = request.form.get('email')
            user = User.query.filter_by(email=email).first()
            if user:
                token = generate_reset_token(email)
                reset_url = url_for('reset_password', token=token, _external=True)
                try:
                    msg = Message(
                        subject='Password Reset Request - School Library System',
                        recipients=[email],
                        html=render_template('reset_password_email.html',
                                             user=user, reset_url=reset_url)
                    )
                    mail.send(msg)
                    _log(db, Log, user.id, 'PASSWORD_RESET_REQUEST',
                         f'Reset requested for {email}', request.remote_addr)
                    db.session.commit()
                    flash('Password reset instructions have been sent to your email.', 'info')
                    return redirect(url_for('login'))
                except Exception as e:
                    logger.error(f"Password reset email error: {e}")
                    flash('Error sending email. Please try again later.', 'danger')
                    return redirect(url_for('forgot_password'))
            # Don't reveal whether email exists
            flash('If your email is registered, you will receive reset instructions.', 'info')
            return redirect(url_for('login'))
        return render_template('forgot_password.html')

    @app.route('/reset-password/<token>', methods=['GET', 'POST'])
    def reset_password(token):
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        email = verify_reset_token(token)
        if not email:
            flash('The password reset link is invalid or has expired.', 'danger')
            return redirect(url_for('forgot_password'))
        if request.method == 'POST':
            password = request.form.get('password')
            confirm = request.form.get('confirm_password')
            if password != confirm:
                flash('Passwords do not match.', 'danger')
                return render_template('reset_password.html', token=token)
            if len(password) < 8:
                flash('Password must be at least 8 characters.', 'danger')
                return render_template('reset_password.html', token=token)
            user = User.query.filter_by(email=email).first()
            if user:
                user.set_password(password)
                _log(db, Log, user.id, 'PASSWORD_RESET',
                     f'Password reset for {email}', request.remote_addr)
                db.session.commit()
                flash('Password reset successfully. You can now log in.', 'success')
                return redirect(url_for('login'))
            flash('User not found.', 'danger')
            return redirect(url_for('forgot_password'))
        return render_template('reset_password.html', token=token)

    # ── Dashboard ─────────────────────────────────────────────────────────────

    @app.route('/dashboard')
    @login_required
    def dashboard():
        stats = {
            'total_books': Book.query.count(),
            'available_books': Book.query.filter_by(is_available=True).count(),
            'borrowed_books': BookBorrowing.query.filter_by(status='borrowed').count(),
            'overdue_books': BookBorrowing.query.filter_by(status='overdue').count(),
            'total_teachers': User.query.filter_by(role='teacher').count(),
        }
        recent_books = Book.query.order_by(Book.date_added.desc()).limit(5).all()
        recent_borrowings = BookBorrowing.query.order_by(
            BookBorrowing.borrowed_date.desc()).limit(5).all()
        return render_template('dashboard.html', stats=stats,
                               recent_books=recent_books,
                               recent_borrowings=recent_borrowings)

    # ── Book management ───────────────────────────────────────────────────────

    @app.route('/books')
    @login_required
    def books():
        q, search, subject, grade, avail = _book_query_from_args(Book, request.args)
        book_list = q.order_by(Book.title).all()
        subjects, grades = _subjects_and_grades(db, Book)
        return render_template('books.html', books=book_list, subjects=subjects,
                               grades=grades, search_query=search,
                               subject_filter=subject, grade_filter=grade,
                               available_filter=avail)

    @app.route('/books/add')
    @librarian_required
    def add_book():
        return redirect(url_for('add_books_bulk'))

    @app.route('/books/add-bulk', methods=['GET', 'POST'])
    @librarian_required
    def add_books_bulk():
        if request.method == 'POST':
            try:
                add_type = request.form.get('add_type', 'single')
                if add_type == 'single':
                    total = int(request.form.get('total_copies', 1))
                    book = Book(
                        title=request.form.get('title'),
                        author=request.form.get('author'),
                        subject=request.form.get('subject'),
                        grade_level=request.form.get('grade_level'),
                        isbn=request.form.get('isbn'),
                        publisher=request.form.get('publisher'),
                        publication_year=request.form.get('publication_year'),
                        edition=request.form.get('edition'),
                        total_copies=total,
                        copies_available=total,
                        is_available=True,
                        shelf_location=request.form.get('shelf_location')
                    )
                    book.book_number = book.generate_book_number()
                    db.session.add(book)
                    book_count = 1
                else:
                    prefix = request.form.get('prefix', '').upper().strip()
                    year = request.form.get('year', str(datetime.now().year))
                    start_num = int(request.form.get('start_num', 1))
                    end_num = int(request.form.get('end_num', 1))
                    if start_num > end_num:
                        flash('Start number must be ≤ end number', 'danger')
                        return redirect(url_for('add_books_bulk'))
                    if end_num - start_num + 1 > 1000:
                        flash('Cannot add more than 1000 books at once', 'danger')
                        return redirect(url_for('add_books_bulk'))
                    base = {
                        'title': request.form.get('bulk_title'),
                        'author': request.form.get('bulk_author'),
                        'subject': request.form.get('bulk_subject'),
                        'grade_level': request.form.get('bulk_grade_level'),
                        'isbn': request.form.get('bulk_isbn'),
                        'publisher': request.form.get('bulk_publisher'),
                        'publication_year': request.form.get('bulk_publication_year'),
                        'edition': request.form.get('bulk_edition'),
                        'total_copies': request.form.get('bulk_total_copies', 1),
                        'shelf_location': request.form.get('bulk_shelf_location'),
                    }
                    books_to_add = Book.create_bulk_books(prefix, year, start_num, end_num, base)
                    for b in books_to_add:
                        db.session.add(b)
                    book_count = len(books_to_add)

                _log(db, Log, current_user.id,
                     'ADD_BOOKS_BULK' if add_type == 'bulk' else 'ADD_BOOK',
                     f'Added {book_count} book(s)', request.remote_addr)
                db.session.commit()
                if add_type == 'single':
                    flash(f'Book added! Number: {book.book_number}', 'success')
                else:
                    flash(f'Added {book_count} books successfully!', 'success')
                return redirect(url_for('books'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding books: {e}', 'danger')
        return render_template('add_books_bulk.html')

    @app.route('/api/check-book-numbers', methods=['POST'])
    @librarian_required
    def check_book_numbers():
        data = request.json
        prefix = data.get('prefix', '').upper().strip()
        year = data.get('year', str(datetime.now().year))
        start_num = int(data.get('start_num', 1))
        end_num = int(data.get('end_num', 1))
        if start_num > end_num:
            return jsonify({'error': 'Invalid range'}), 400
        requested = [f"{prefix}/{year}/{str(n).zfill(3)}" for n in range(start_num, end_num + 1)]
        existing = [b.book_number for b in Book.query.filter(Book.book_number.in_(requested)).all()]
        return jsonify({'requested': len(requested), 'existing': len(existing),
                        'available': len(requested) - len(existing),
                        'existing_numbers': existing})

    @app.route('/books/edit/<int:book_id>', methods=['GET', 'POST'])
    @librarian_required
    def edit_book(book_id):
        book = Book.query.get_or_404(book_id)
        if request.method == 'POST':
            try:
                book.title = request.form.get('title')
                book.author = request.form.get('author')
                book.subject = request.form.get('subject')
                book.grade_level = request.form.get('grade_level')
                book.isbn = request.form.get('isbn')
                book.publisher = request.form.get('publisher')
                book.publication_year = request.form.get('publication_year')
                book.edition = request.form.get('edition')
                book.total_copies = int(request.form.get('total_copies', 1))
                book.shelf_location = request.form.get('shelf_location')
                borrowed_count = BookBorrowing.query.filter_by(
                    book_id=book.id, status='borrowed').count()
                book.copies_available = book.total_copies - borrowed_count
                book.is_available = book.copies_available > 0
                _log(db, Log, current_user.id, 'EDIT_BOOK',
                     f'Edited: {book.title} ({book.book_number})', request.remote_addr)
                db.session.commit()
                flash('Book updated successfully!', 'success')
                return redirect(url_for('books'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating book: {e}', 'danger')
        return render_template('edit_book.html', book=book)

    @app.route('/books/bulk-edit', methods=['GET', 'POST'])
    @librarian_required
    def bulk_edit_books():
        if request.method == 'POST':
            book_ids = request.form.getlist('book_ids')
            if not book_ids:
                flash('No books selected', 'warning')
                return redirect(url_for('books'))

            if 'apply_edits' in request.form:
                try:
                    editable = ['title', 'author', 'subject', 'grade_level', 'isbn',
                                'publisher', 'publication_year', 'edition', 'shelf_location']
                    to_update = {f: request.form.get(f) for f in editable
                                 if request.form.get(f'update_{f}') == 'on'}
                    if not to_update:
                        flash('No fields selected for update', 'warning')
                        return redirect(url_for('bulk_edit_books'))

                    updated = 0
                    for bid in book_ids:
                        book = Book.query.get(int(bid))
                        if book:
                            for field, value in to_update.items():
                                if field == 'publication_year' and value:
                                    setattr(book, field, Book._parse_year(value))
                                elif value:
                                    setattr(book, field, value)
                            updated += 1

                    _log(db, Log, current_user.id, 'BULK_EDIT_BOOKS',
                         f'Bulk edited {updated} books. Fields: {", ".join(to_update)}',
                         request.remote_addr)
                    db.session.commit()
                    flash(f'Updated {updated} book(s)', 'success')
                    return redirect(url_for('bulk_edit_books'))
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error: {e}', 'danger')

            selected_books = Book.query.filter(Book.id.in_(book_ids)).all()
            if not selected_books:
                flash('No valid books found', 'warning')
                return redirect(url_for('books'))
            common = {}
            for field in ['subject', 'grade_level', 'shelf_location']:
                vals = {getattr(b, field) for b in selected_books if getattr(b, field)}
                common[field] = vals.pop() if len(vals) == 1 else ''
            return render_template('bulk_edit_books.html', books=selected_books,
                                   book_ids=book_ids, common_values=common,
                                   total_books=len(selected_books))

        q, search, subject, grade, avail = _book_query_from_args(Book, request.args)
        book_list = q.order_by(Book.title).all()
        subjects, grades = _subjects_and_grades(db, Book)
        return render_template('select_books_for_edit.html', books=book_list,
                               subjects=subjects, grades=grades, search_query=search,
                               subject_filter=subject, grade_filter=grade,
                               available_filter=avail)

    @app.route('/books/delete/<int:book_id>', methods=['POST'])
    @librarian_required
    def delete_book(book_id):
        book = Book.query.get_or_404(book_id)
        if BookBorrowing.query.filter_by(book_id=book_id, status='borrowed').count():
            flash('Cannot delete a book that is currently borrowed', 'danger')
            return redirect(url_for('books'))
        try:
            BookBorrowing.query.filter_by(book_id=book_id).delete()
            _log(db, Log, current_user.id, 'DELETE_BOOK',
                 f'Deleted: {book.title} ({book.book_number})', request.remote_addr)
            db.session.delete(book)
            db.session.commit()
            flash('Book deleted successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting book: {e}', 'danger')
        return redirect(url_for('books'))

    @app.route('/books/bulk-delete', methods=['GET', 'POST'])
    @librarian_required
    def bulk_delete_books():
        if request.method == 'POST':
            book_ids = request.form.getlist('book_ids')
            if not book_ids:
                flash('No books selected', 'warning')
                return redirect(url_for('books'))
            try:
                deleted, failed, errors = 0, 0, []
                for bid in book_ids:
                    try:
                        book = Book.query.get(int(bid))
                        if not book:
                            failed += 1
                            errors.append(f"ID {bid} not found")
                            continue
                        if BookBorrowing.query.filter_by(book_id=book.id, status='borrowed').count():
                            failed += 1
                            errors.append(f"'{book.title}' is currently borrowed")
                            continue
                        BookBorrowing.query.filter_by(book_id=book.id).delete()
                        db.session.delete(book)
                        deleted += 1
                    except Exception as e:
                        failed += 1
                        errors.append(f"ID {bid}: {e}")

                _log(db, Log, current_user.id, 'BULK_DELETE_BOOKS',
                     f'Deleted {deleted} books in bulk', request.remote_addr)
                db.session.commit()
                if deleted:
                    flash(f'Deleted {deleted} book(s)', 'success')
                if failed:
                    msg = f'Failed to delete {failed} book(s):<br>' + '<br>'.join(errors[:3])
                    if len(errors) > 3:
                        msg += f'<br>... and {len(errors)-3} more'
                    flash(msg, 'warning')
                return redirect(url_for('books'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')
                return redirect(url_for('books'))

        q, search, subject, grade, avail = _book_query_from_args(Book, request.args)
        book_list = q.order_by(Book.title).all()
        subjects, grades = _subjects_and_grades(db, Book)
        return render_template('bulk_delete_books.html', books=book_list,
                               subjects=subjects, grades=grades, search_query=search,
                               subject_filter=subject, grade_filter=grade,
                               available_filter=avail)

    @app.route('/books/delete-check', methods=['POST'])
    @librarian_required
    def check_books_for_deletion():
        data = request.json
        book_ids = data.get('book_ids', [])
        if not book_ids:
            return jsonify({'error': 'No books selected'}), 400
        result = {'total_selected': len(book_ids), 'can_delete': 0,
                  'cannot_delete': 0, 'books': []}
        for bid in book_ids:
            try:
                book = Book.query.get(int(bid))
                if not book:
                    result['books'].append({'id': bid, 'title': 'Not Found',
                                            'book_number': 'N/A', 'can_delete': False,
                                            'reason': 'Book not found'})
                    result['cannot_delete'] += 1
                    continue
                active = BookBorrowing.query.filter_by(book_id=book.id, status='borrowed').count()
                if active:
                    result['books'].append({'id': bid, 'title': book.title,
                                            'book_number': book.book_number, 'can_delete': False,
                                            'reason': f'Currently borrowed ({active} copies)'})
                    result['cannot_delete'] += 1
                else:
                    result['books'].append({'id': bid, 'title': book.title,
                                            'book_number': book.book_number, 'can_delete': True,
                                            'reason': 'Can be deleted'})
                    result['can_delete'] += 1
            except Exception as e:
                result['books'].append({'id': bid, 'title': 'Error', 'book_number': 'N/A',
                                        'can_delete': False, 'reason': str(e)})
                result['cannot_delete'] += 1
        return jsonify(result)

    # ── Borrowing ─────────────────────────────────────────────────────────────

    @app.route('/borrowing')
    @login_required
    def borrowing():
        user_id = request.args.get('user_id', type=int)
        status_filter = request.args.get('status', '')
        q = BookBorrowing.query
        if user_id:
            q = q.filter_by(user_id=user_id)
        if status_filter:
            q = q.filter_by(status=status_filter)
        borrowings = q.order_by(BookBorrowing.borrowed_date.desc()).all()

        # Calculate fines — single commit after the loop
        updated = False
        for b in borrowings:
            if b.status == 'borrowed' and utcnow() > b.due_date.replace(tzinfo=timezone.utc):
                b.calculate_fine()
                updated = True
        if updated:
            db.session.commit()

        teachers = User.query.filter_by(role='teacher').all()
        return render_template('borrowing.html', borrowings=borrowings,
                               teachers=teachers, user_id=user_id,
                               status_filter=status_filter)

    @app.route('/borrow/teacher/<int:teacher_id>')
    @librarian_required
    def teacher_borrowings(teacher_id):
        teacher = User.query.get_or_404(teacher_id)
        borrowings = BookBorrowing.query.filter_by(
            user_id=teacher_id).order_by(BookBorrowing.borrowed_date.desc()).all()
        return render_template('teacher_borrowings.html', teacher=teacher,
                               borrowings=borrowings)

    @app.route('/borrow/bulk', methods=['GET', 'POST'])
    @librarian_required
    def bulk_borrow():
        if request.method == 'POST':
            teacher_id = request.form.get('teacher_id')
            book_ids = request.form.getlist('book_ids')
            due_date_str = request.form.get('due_date')

            # Validate required fields
            if not teacher_id:
                flash('Please select a teacher', 'danger')
                return redirect(url_for('bulk_borrow'))
            if not due_date_str:
                flash('Please select a due date', 'danger')
                return redirect(url_for('bulk_borrow'))
            if not book_ids:
                flash('Please select at least one book', 'danger')
                return redirect(url_for('bulk_borrow'))

            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
                teacher = db.session.get(User, teacher_id)
                if not teacher:
                    flash('Teacher not found', 'danger')
                    return redirect(url_for('bulk_borrow'))
                borrowed_books = []
                for bid in book_ids:
                    book = db.session.get(Book, int(bid))
                    if book and book.copies_available > 0:
                        borrow = BookBorrowing(book_id=book.id, user_id=teacher.id,
                                              due_date=due_date, status='borrowed')
                        book.copies_available -= 1
                        if book.copies_available == 0:
                            book.is_available = False
                        db.session.add(borrow)
                        borrowed_books.append({'book': book, 'borrowing': borrow})
                db.session.commit()

                receipt_number = None
                if borrowed_books and teacher.email:
                    books_list = [item['book'] for item in borrowed_books]
                    _, receipt_number = send_borrow_notification(
                        borrowing=borrowed_books[0]['borrowing'], book=None,
                        borrower=teacher, is_bulk=True, books_list=books_list)

                _log(db, Log, current_user.id, 'BULK_BORROW',
                     f'{teacher.full_name} borrowed {len(borrowed_books)} books',
                     request.remote_addr)
                db.session.commit()

                if teacher.email:
                    if receipt_number:
                        flash(f'{len(borrowed_books)} books borrowed! Receipt #{receipt_number} sent to {teacher.email}', 'success')
                    else:
                        flash(f'{len(borrowed_books)} books borrowed! (Email failed)', 'warning')
                else:
                    flash(f'{len(borrowed_books)} books borrowed! (No email on file)', 'info')
                return redirect(url_for('borrowing'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')

        teachers = User.query.filter_by(role='teacher').all()
        # Pass ALL books — template marks unavailable ones as greyed out/disabled
        # Sorting: available first, then unavailable
        all_books = Book.query.order_by(
            Book.is_available.desc(), Book.subject, Book.book_number
        ).all()
        subjects, grades = _subjects_and_grades(db, Book)
        return render_template('bulk_borrow.html', teachers=teachers,
                               books=all_books, subjects=subjects, grades=grades)

    @app.route('/return/bulk', methods=['POST'])
    @librarian_required
    def bulk_return():
        borrowing_ids = request.form.getlist('borrowing_ids')
        try:
            returned = 0
            for bid in borrowing_ids:
                b = BookBorrowing.query.get(int(bid))
                if b and b.status == 'borrowed':
                    b.status = 'returned'
                    b.returned_date = utcnow()
                    book = Book.query.get(b.book_id)
                    if book:
                        book.copies_available += 1
                        book.is_available = True
                    returned += 1
            _log(db, Log, current_user.id, 'BULK_RETURN',
                 f'Returned {returned} books', request.remote_addr)
            db.session.commit()
            flash(f'{returned} books returned successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
        return redirect(url_for('borrowing'))

    @app.route('/return/select-teacher')
    @librarian_required
    def select_teacher_for_return():
        teachers = User.query.filter_by(role='teacher').order_by(User.full_name).all()
        now = utcnow()
        stats = []
        for t in teachers:
            borrowed = BookBorrowing.query.filter_by(user_id=t.id, status='borrowed').count()
            overdue = BookBorrowing.query.filter(
                BookBorrowing.user_id == t.id,
                BookBorrowing.status == 'borrowed',
                BookBorrowing.due_date < now
            ).count()
            stats.append({'teacher': t, 'borrowed_count': borrowed, 'overdue_count': overdue})
        return render_template('select_teacher_return.html', teacher_stats=stats)

    @app.route('/return/teacher/<int:teacher_id>', methods=['GET', 'POST'])
    @librarian_required
    def return_teacher_books(teacher_id):
        teacher = User.query.get_or_404(teacher_id)
        if request.method == 'POST':
            borrowing_ids = request.form.getlist('borrowing_ids')
            return_date_str = request.form.get('return_date', utcnow().strftime('%Y-%m-%d'))
            try:
                returned_count, fine_collected, returned_books_list = 0, 0, []
                for bid in borrowing_ids:
                    b = db.session.get(BookBorrowing, int(bid))
                    if b and b.status == 'borrowed':
                        book = db.session.get(Book, b.book_id)
                        old_fine = b.fine_amount
                        b.status = 'returned'
                        b.returned_date = datetime.strptime(return_date_str, '%Y-%m-%d')
                        fine_paid = request.form.get(f'fine_paid_{bid}') == 'on'
                        if fine_paid:
                            b.fine_paid = True
                            fine_collected += b.fine_amount
                        if book:
                            book.copies_available += 1
                            book.is_available = True
                        returned_books_list.append({'book': book, 'borrowing': b,
                                                    'fine': old_fine, 'fine_paid': fine_paid})
                        returned_count += 1
                db.session.commit()

                receipt_number = None
                if returned_books_list and teacher.email:
                    total_fine = sum(i['fine'] for i in returned_books_list)
                    fines_paid = sum(i['fine'] for i in returned_books_list if i['fine_paid'])
                    _, receipt_number = send_return_notification(
                        borrowing=None, book=None, borrower=teacher, is_bulk=True,
                        returned_books_list=returned_books_list,
                        total_fine=total_fine, fines_paid=fines_paid)

                _log(db, Log, current_user.id, 'TEACHER_RETURN',
                     f'Returned {returned_count} books for {teacher.full_name}',
                     request.remote_addr)
                db.session.commit()

                if teacher.email:
                    if receipt_number:
                        flash(f'Returned {returned_count} book(s) for {teacher.full_name}. Receipt #{receipt_number} sent.', 'success')
                    else:
                        flash(f'Returned {returned_count} book(s) (Email failed)', 'warning')
                else:
                    flash(f'Returned {returned_count} book(s) for {teacher.full_name}', 'success')
                if fine_collected > 0:
                    flash(f'Fine collected: ${fine_collected:.2f}', 'info')
                return redirect(url_for('select_teacher_for_return'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')

        borrowed_books = BookBorrowing.query.filter_by(
            user_id=teacher_id, status='borrowed').order_by(BookBorrowing.due_date).all()
        now = utcnow()
        for b in borrowed_books:
            due = b.due_date
            if due.tzinfo is None:
                due = due.replace(tzinfo=timezone.utc)
            if now > due:
                b.calculate_fine()
        return render_template('return_teacher_books.html', teacher=teacher,
                               borrowed_books=borrowed_books)

    @app.route('/return/single/<int:borrowing_id>', methods=['POST'])
    @librarian_required
    def return_single_book(borrowing_id):
        b = BookBorrowing.query.get_or_404(borrowing_id)
        if b.status != 'borrowed':
            flash('This book is not currently borrowed', 'danger')
            return redirect(request.referrer or url_for('borrowing'))
        try:
            book = Book.query.get(b.book_id)
            teacher = b.borrower
            b.status = 'returned'
            b.returned_date = utcnow()
            if book:
                book.copies_available += 1
                book.is_available = True
            db.session.commit()
            if teacher and teacher.email:
                send_return_notification(borrowing=b, book=book, borrower=teacher)
            _log(db, Log, current_user.id, 'SINGLE_RETURN',
                 f'Returned: {book.title if book else "?"} for {teacher.full_name if teacher else "?"}',
                 request.remote_addr)
            db.session.commit()
            email_status = " (Email sent)" if teacher and teacher.email else " (No email on file)"
            flash(f'Book returned successfully!{email_status}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
        return redirect(request.referrer or url_for('borrowing'))

    # ── Reports ───────────────────────────────────────────────────────────────

    @app.route('/reports')
    @login_required
    def reports():
        teachers = User.query.filter_by(role='teacher').all()
        return render_template('reports.html', teachers=teachers)

    @app.route('/reports/overdue')
    @librarian_required
    def overdue_report():
        data = BookBorrowing.query.filter_by(status='overdue').all()
        fmt = request.args.get('format', 'html')
        if fmt == 'csv':
            return generate_csv_report(data, 'overdue')
        elif fmt == 'excel':
            return generate_excel_report(data, 'overdue')
        elif fmt == 'pdf':
            return generate_pdf_report(data, 'overdue')
        return render_template('overdue_report.html', borrowings=data)

    @app.route('/reports/teacher/<int:teacher_id>')
    @librarian_required
    def teacher_report(teacher_id):
        teacher = User.query.get_or_404(teacher_id)
        data = BookBorrowing.query.filter_by(user_id=teacher_id).all()
        fmt = request.args.get('format', 'html')
        if fmt != 'html' and not data:
            flash(f'No borrowing records for {teacher.full_name}', 'warning')
            return redirect(url_for('reports'))
        if fmt == 'csv':
            return generate_csv_report(data, f'teacher_{teacher.username}')
        elif fmt == 'excel':
            return generate_excel_report(data, f'teacher_{teacher.username}')
        elif fmt == 'pdf':
            return generate_pdf_report(data, f'teacher_{teacher.username}')
        return render_template('teacher_report.html', teacher=teacher, borrowings=data)

    @app.route('/reports/inventory')
    @librarian_required
    def inventory_report():
        books = Book.query.all()
        fmt = request.args.get('format', 'html')
        if fmt == 'csv':
            return generate_csv_report(books, 'inventory')
        elif fmt == 'excel':
            return generate_excel_report(books, 'inventory')
        elif fmt == 'pdf':
            return generate_pdf_report(books, 'inventory')
        return render_template('inventory_report.html', books=books)

    # ── Staff management ──────────────────────────────────────────────────────

    @app.route('/staff')
    @admin_required
    def staff():
        return render_template('staff.html', staff=User.query.all())

    @app.route('/staff/add', methods=['GET', 'POST'])
    @admin_required
    def add_staff():
        if request.method == 'POST':
            try:
                member = User(
                    username=request.form.get('username'),
                    email=request.form.get('email'),
                    full_name=request.form.get('full_name'),
                    role=request.form.get('role'),
                    department=request.form.get('department')
                )
                pw = request.form.get('password')
                if pw:
                    member.set_password(pw)
                db.session.add(member)
                _log(db, Log, current_user.id, 'ADD_STAFF',
                     f'Added: {member.full_name} ({member.role})', request.remote_addr)
                db.session.commit()
                flash('Staff member added!', 'success')
                return redirect(url_for('staff'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')
        return render_template('add_staff.html')

    @app.route('/staff/toggle/<int:staff_id>', methods=['POST'])
    @admin_required
    def toggle_staff_status(staff_id):
        member = User.query.get_or_404(staff_id)
        if member.id == current_user.id:
            flash('You cannot change your own status', 'danger')
            return redirect(url_for('staff'))
        try:
            member.is_active = not member.is_active
            status = 'activated' if member.is_active else 'deactivated'
            _log(db, Log, current_user.id, 'TOGGLE_STAFF_STATUS',
                 f'{status}: {member.full_name}', request.remote_addr)
            db.session.commit()
            flash(f'Staff member {status}!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
        return redirect(url_for('staff'))

    @app.route('/staff/delete/<int:staff_id>', methods=['POST'])
    @admin_required
    def delete_staff(staff_id):
        member = User.query.get_or_404(staff_id)
        if member.id == current_user.id:
            flash('You cannot delete your own account', 'danger')
            return redirect(url_for('staff'))
        if BookBorrowing.query.filter_by(user_id=staff_id, status='borrowed').count():
            flash('Cannot delete staff with borrowed books', 'danger')
            return redirect(url_for('staff'))
        try:
            _log(db, Log, current_user.id, 'DELETE_STAFF',
                 f'Deleted: {member.full_name}', request.remote_addr)
            db.session.delete(member)
            db.session.commit()
            flash('Staff member deleted!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
        return redirect(url_for('staff'))

    @app.route('/staff/edit/<int:staff_id>', methods=['GET', 'POST'])
    @admin_required
    def edit_staff(staff_id):
        member = User.query.get_or_404(staff_id)
        if member.id == current_user.id:
            flash('Edit your own profile from the Profile page', 'warning')
            return redirect(url_for('profile'))
        if request.method == 'POST':
            try:
                member.full_name = request.form.get('full_name')
                member.email = request.form.get('email')
                member.username = request.form.get('username')
                member.role = request.form.get('role')
                member.department = request.form.get('department')
                member.is_active = 'is_active' in request.form

                if request.form.get('reset_password') == 'on':
                    # Generate temp password and email it — don't flash it
                    temp_pw = ''.join(
                        secrets_module.choice(string.ascii_letters + string.digits)
                        for _ in range(12)
                    )
                    member.set_password(temp_pw)
                    try:
                        msg = Message(
                            subject='Your Library Account Password Has Been Reset',
                            recipients=[member.email],
                            body=f"Hello {member.full_name},\n\n"
                                 f"Your temporary password is: {temp_pw}\n\n"
                                 f"Please log in and change it immediately."
                        )
                        mail.send(msg)
                        flash('Password reset. Temporary password emailed to the staff member.', 'info')
                    except Exception:
                        flash(f'Password reset. Temporary password (email failed): {temp_pw}', 'warning')

                _log(db, Log, current_user.id, 'EDIT_STAFF',
                     f'Edited: {member.full_name}', request.remote_addr)
                db.session.commit()
                flash(f'{member.full_name} updated!', 'success')
                return redirect(url_for('staff'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')

        total_b = BookBorrowing.query.filter_by(user_id=member.id).count()
        active_b = BookBorrowing.query.filter_by(user_id=member.id, status='borrowed').count()
        overdue_b = BookBorrowing.query.filter(
            BookBorrowing.user_id == member.id,
            BookBorrowing.status == 'borrowed',
            BookBorrowing.due_date < utcnow()
        ).count()
        returned_b = BookBorrowing.query.filter_by(user_id=member.id, status='returned').count()
        return render_template('edit_staff.html', staff=member,
                               total_borrowings=total_b, active_borrowings=active_b,
                               overdue_borrowings=overdue_b, returned_borrowings=returned_b)

    # ── Profile ───────────────────────────────────────────────────────────────

    @app.route('/profile')
    @login_required
    def profile():
        borrowings = BookBorrowing.query.filter_by(
            user_id=current_user.id).order_by(BookBorrowing.borrowed_date.desc()).all()
        return render_template('profile.html', borrowings=borrowings)

    # ── Import / Export ───────────────────────────────────────────────────────

    @app.route('/import-export')
    @librarian_required
    def import_export():
        return render_template('import_export.html')

    @app.route('/books/import', methods=['GET', 'POST'])
    @librarian_required
    def import_books():
        if request.method == 'POST':
            if 'file' not in request.files or request.files['file'].filename == '':
                flash('No file selected', 'danger')
                return redirect(url_for('import_books'))

            file = request.files['file']
            ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
            skip_dup = request.form.get('skip_duplicates') == 'on'
            overwrite = request.form.get('overwrite_existing') == 'on'

            created, updated, skipped, errors = 0, 0, 0, []

            try:
                rows = []
                if ext == 'csv':
                    stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
                    rows = list(csv.DictReader(stream))

                elif ext in ('xlsx', 'xls'):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()))
                        ws = wb.active
                        headers = [str(c.value).strip().lower() if c.value else ''
                                   for c in ws[1]]
                        for row in ws.iter_rows(min_row=2):
                            rows.append({headers[j]: cell.value
                                         for j, cell in enumerate(row) if headers[j]})
                    except ImportError:
                        flash('Excel import requires openpyxl: pip install openpyxl', 'warning')
                        return redirect(url_for('import_books'))

                elif ext == 'json':
                    data = json.loads(file.read().decode('utf-8'))
                    if not isinstance(data, list):
                        flash('JSON must be an array of books', 'danger')
                        return redirect(url_for('import_books'))
                    rows = data
                else:
                    flash('Unsupported format. Use CSV, Excel, or JSON.', 'danger')
                    return redirect(url_for('import_books'))

                for i, row in enumerate(rows, 1):
                    outcome, err = _upsert_book(Book, db, row, skip_dup, overwrite)
                    if outcome == 'created':
                        created += 1
                    elif outcome == 'updated':
                        updated += 1
                    else:
                        skipped += 1
                        if err:
                            errors.append(f"Row {i}: {err}")

                if created or updated:
                    _log(db, Log, current_user.id, 'IMPORT_BOOKS',
                         f'Imported {created} new, updated {updated}', request.remote_addr)
                    db.session.commit()
                    parts = []
                    if created:
                        parts.append(f'{created} new book(s) imported')
                    if updated:
                        parts.append(f'{updated} book(s) updated')
                    flash(' | '.join(parts), 'success')
                    if skipped:
                        flash(f'{skipped} row(s) skipped', 'warning')
                    if errors:
                        flash('Errors:<br>' + '<br>'.join(errors[:5]) +
                              (f'<br>...and {len(errors)-5} more' if len(errors) > 5 else ''),
                              'warning')
                else:
                    flash('No books imported. Check your file.', 'warning')

                return redirect(url_for('books'))
            except Exception as e:
                db.session.rollback()
                flash(f'Import error: {e}', 'danger')
                return redirect(url_for('import_books'))

        return render_template('import_books.html')

    @app.route('/books/export')
    @librarian_required
    def export_books():
        fmt = request.args.get('format', 'csv')
        filter_type = request.args.get('filter', 'all')
        q = Book.query
        if filter_type == 'available':
            q = q.filter_by(is_available=True)
        elif filter_type == 'borrowed':
            q = q.filter_by(is_available=False)
        elif filter_type == 'subject':
            subj = request.args.get('subject', '')
            if subj:
                q = q.filter_by(subject=subj)
        books = q.order_by(Book.subject, Book.title).all()
        if not books:
            flash('No books found for export', 'warning')
            return redirect(url_for('import_export'))
        if fmt == 'csv':
            return export_books_csv(books)
        elif fmt == 'excel':
            return export_books_excel(books)
        elif fmt == 'json':
            return export_books_json(books)
        elif fmt == 'pdf':
            return export_books_pdf(books)
        flash('Invalid export format', 'danger')
        return redirect(url_for('import_export'))

    @app.route('/borrowings/export')
    @librarian_required
    def export_borrowings():
        fmt = request.args.get('format', 'csv')
        status_filter = request.args.get('status', '')
        teacher_id = request.args.get('teacher_id', type=int)
        q = BookBorrowing.query
        if status_filter:
            q = q.filter_by(status=status_filter)
        if teacher_id:
            q = q.filter_by(user_id=teacher_id)
        data = q.order_by(BookBorrowing.borrowed_date.desc()).all()
        if not data:
            flash('No borrowing records found', 'warning')
            return redirect(url_for('import_export'))
        if fmt == 'csv':
            return export_borrowings_csv(data)
        elif fmt == 'excel':
            return export_borrowing_excel(data)
        elif fmt == 'pdf':
            return export_borrowings_pdf(data)
        flash('Invalid format', 'danger')
        return redirect(url_for('import_export'))

    @app.route('/books/bulk-return', methods=['GET', 'POST'])
    @librarian_required
    def bulk_return_from_file():
        if request.method == 'POST':
            if 'file' not in request.files or request.files['file'].filename == '':
                flash('No file selected', 'danger')
                return redirect(url_for('bulk_return_from_file'))
            file = request.files['file']
            try:
                returned, not_found, already_returned = 0, 0, 0
                stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
                for row in csv.DictReader(stream):
                    bn = row.get('book_number', '').strip()
                    if not bn:
                        continue
                    book = Book.query.filter_by(book_number=bn).first()
                    if not book:
                        not_found += 1
                        continue
                    b = BookBorrowing.query.filter_by(book_id=book.id, status='borrowed').first()
                    if not b:
                        already_returned += 1
                        continue
                    b.status = 'returned'
                    b.returned_date = utcnow()
                    book.copies_available += 1
                    book.is_available = True
                    returned += 1

                _log(db, Log, current_user.id, 'BULK_RETURN_FILE',
                     f'Returned {returned} books from file', request.remote_addr)
                db.session.commit()
                flash(f'Returned {returned} books from file', 'success')
                if not_found:
                    flash(f'{not_found} books not found', 'warning')
                if already_returned:
                    flash(f'{already_returned} already returned', 'info')
                return redirect(url_for('borrowing'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')
        return render_template('bulk_return_file.html')

    @app.route('/download-template')
    @librarian_required
    def download_template():
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(['book_number', 'title', 'author', 'subject', 'grade_level',
                         'isbn', 'publisher', 'publication_year', 'edition',
                         'total_copies', 'copies_available', 'shelf_location'])
        examples = [
            ['ENG/2026/001', 'English Literature', 'John Smith', 'English', 'Grade 10',
             '9781234567890', 'Education Press', '2023', 'First Edition', '5', '5', 'Shelf A-12'],
            ['MAT/2026/001', 'Mathematics Grade 10', 'Jane Doe', 'Mathematics', 'Grade 10',
             '9780987654321', 'Math Publishers', '2022', 'Second Edition', '3', '3', 'Shelf B-05'],
        ]
        for row in examples:
            writer.writerow(row)
        si.seek(0)
        return send_file(io.BytesIO(si.getvalue().encode('utf-8')),
                         mimetype='text/csv', as_attachment=True,
                         download_name='book_import_template.csv')

    @app.route('/admin/fix-book-availability')
    @admin_required
    def fix_book_availability():
        """Fix is_available flag for all books based on actual borrowed copies."""
        try:
            books = Book.query.all()
            fixed = 0
            for book in books:
                borrowed = BookBorrowing.query.filter_by(
                    book_id=book.id, status='borrowed').count()
                correct_available = book.total_copies - borrowed
                correct_is_available = correct_available > 0
                if (book.copies_available != correct_available or
                        book.is_available != correct_is_available):
                    book.copies_available = max(0, correct_available)
                    book.is_available = correct_is_available
                    fixed += 1
            db.session.commit()
            flash(f'Fixed availability for {fixed} book(s). All books now showing correctly.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
        return redirect(url_for('books'))

    @app.route('/network-access')
    @login_required
    def network_access():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            server_ip = s.getsockname()[0]
            s.close()
        except Exception:
            server_ip = "Unable to determine IP"
        return render_template('network_access.html', server_ip=server_ip,
                               port=5000, computer_name=socket.gethostname())

    # ── Export helpers ────────────────────────────────────────────────────────

    _BOOK_HEADERS = ['Book Number', 'Title', 'Author', 'Subject', 'Grade Level',
                     'ISBN', 'Publisher', 'Publication Year', 'Edition',
                     'Total Copies', 'Available Copies', 'Shelf Location', 'Date Added', 'Status']

    def _book_row(b):
        return [b.book_number, b.title, b.author, b.subject, b.grade_level or '',
                b.isbn or '', b.publisher or '', b.publication_year or '',
                b.edition or '', b.total_copies, b.copies_available,
                b.shelf_location or '',
                b.date_added.strftime('%Y-%m-%d') if b.date_added else '',
                'Available' if b.is_available else 'Borrowed']

    _BORROW_HEADERS = ['Book Number', 'Book Title', 'Borrower Name', 'Borrower Username',
                       'Borrowed Date', 'Due Date', 'Returned Date', 'Status', 'Fine Amount', 'Fine Paid']

    def _borrow_row(b):
        return [b.book.book_number if b.book else '',
                b.book.title if b.book else '',
                b.borrower.full_name if b.borrower else '',
                b.borrower.username if b.borrower else '',
                b.borrowed_date.strftime('%Y-%m-%d') if b.borrowed_date else '',
                b.due_date.strftime('%Y-%m-%d') if b.due_date else '',
                b.returned_date.strftime('%Y-%m-%d') if b.returned_date else '',
                b.status, b.fine_amount, 'Yes' if b.fine_paid else 'No']

    def export_books_csv(books):
        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(_BOOK_HEADERS)
        for b in books:
            w.writerow(_book_row(b))
        si.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(io.BytesIO(si.getvalue().encode('utf-8')),
                         mimetype='text/csv', as_attachment=True,
                         download_name=f'books_export_{ts}.csv')

    def export_borrowings_csv(borrowings):
        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(_BORROW_HEADERS)
        for b in borrowings:
            w.writerow(_borrow_row(b))
        si.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(io.BytesIO(si.getvalue().encode('utf-8')),
                         mimetype='text/csv', as_attachment=True,
                         download_name=f'borrowings_export_{ts}.csv')

    def export_books_excel(books):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Books Export"
            for col, h in enumerate(_BOOK_HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for row, b in enumerate(books, 2):
                for col, val in enumerate(_book_row(b), 1):
                    ws.cell(row=row, column=col, value=val)
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            return send_file(out,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'books_export_{ts}.xlsx')
        except ImportError:
            flash('Excel export requires openpyxl: pip install openpyxl', 'warning')
            return redirect(url_for('import_export'))

    def export_borrowing_excel(borrowings):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            STATUS_COLORS = {
                'overdue': 'FFB6C1', 'borrowed': 'FFE4B5', 'returned': 'C0F0C0'
            }
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Borrowing History"
            hfill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col, h in enumerate(_BORROW_HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')
            for row, b in enumerate(borrowings, 2):
                for col, val in enumerate(_borrow_row(b), 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    color = STATUS_COLORS.get(b.status)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            return send_file(out,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'borrowing_history_{ts}.xlsx')
        except ImportError:
            flash('Excel export requires openpyxl: pip install openpyxl', 'warning')
            return redirect(url_for('import_export'))

    def export_books_json(books):
        data = [b.to_dict() for b in books]
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(io.BytesIO(json.dumps(data, indent=2, ensure_ascii=False).encode('utf-8')),
                         mimetype='application/json', as_attachment=True,
                         download_name=f'books_export_{ts}.json')

    def export_borrowings_pdf(borrowings):
        return generate_pdf_report(borrowings, 'borrowings')

    def export_books_pdf(books):
        return generate_pdf_report(books, 'inventory')

    # ── Report helpers ────────────────────────────────────────────────────────

    def generate_csv_report(data, report_name):
        si = io.StringIO()
        if not data:
            return send_file(io.BytesIO(b''), mimetype='text/csv', as_attachment=True,
                             download_name=f'{report_name}_empty.csv')
        w = csv.writer(si)
        if isinstance(data[0], BookBorrowing):
            w.writerow(['Book Title', 'Book Number', 'Borrower', 'Borrowed Date',
                        'Due Date', 'Returned Date', 'Status', 'Fine'])
            for item in data:
                w.writerow([item.book.title if item.book else 'N/A',
                             item.book.book_number if item.book else 'N/A',
                             item.borrower.full_name if item.borrower else 'N/A',
                             item.borrowed_date.strftime('%Y-%m-%d') if item.borrowed_date else '',
                             item.due_date.strftime('%Y-%m-%d') if item.due_date else '',
                             item.returned_date.strftime('%Y-%m-%d') if item.returned_date else '',
                             item.status, item.fine_amount])
        elif isinstance(data[0], Book):
            w.writerow(['Book Number', 'Title', 'Author', 'Subject', 'Grade',
                        'Available', 'Total', 'Shelf Location'])
            for item in data:
                w.writerow([item.book_number, item.title, item.author, item.subject,
                             item.grade_level, item.copies_available,
                             item.total_copies, item.shelf_location])
        si.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(io.BytesIO(si.getvalue().encode('utf-8')), mimetype='text/csv',
                         as_attachment=True, download_name=f'{report_name}_report_{ts}.csv')

    def generate_excel_report(data, report_name):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            if not data:
                ws.append(['No data available'])
            elif isinstance(data[0], BookBorrowing):
                ws.append(['Book Title', 'Book Number', 'Borrower', 'Borrowed Date',
                           'Due Date', 'Returned Date', 'Status', 'Fine'])
                for item in data:
                    ws.append([item.book.title if item.book else 'N/A',
                                item.book.book_number if item.book else 'N/A',
                                item.borrower.full_name if item.borrower else 'N/A',
                                item.borrowed_date.strftime('%Y-%m-%d') if item.borrowed_date else '',
                                item.due_date.strftime('%Y-%m-%d') if item.due_date else '',
                                item.returned_date.strftime('%Y-%m-%d') if item.returned_date else '',
                                item.status, item.fine_amount])
            elif isinstance(data[0], Book):
                ws.append(['Book Number', 'Title', 'Author', 'Subject', 'Grade',
                           'Available', 'Total', 'Shelf Location'])
                for item in data:
                    ws.append([item.book_number, item.title, item.author, item.subject,
                                item.grade_level, item.copies_available,
                                item.total_copies, item.shelf_location])
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            return send_file(out,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'{report_name}_report_{ts}.xlsx')
        except ImportError:
            flash('Excel export requires openpyxl: pip install openpyxl', 'warning')
            return redirect(url_for('reports'))

    def generate_pdf_report(data, report_name):
        buf = io.BytesIO()
        p = canvas.Canvas(buf, pagesize=letter)
        _, height = letter
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, height - 50, f"Library Report: {report_name}")
        p.setFont("Helvetica", 10)
        p.drawString(100, height - 70, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        if not data:
            p.drawString(100, height - 120, "No records found.")
            p.save()
            buf.seek(0)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            return send_file(buf, mimetype='application/pdf', as_attachment=True,
                             download_name=f'{report_name}_report_{ts}.pdf')
        y = height - 100
        if isinstance(data[0], BookBorrowing):
            p.setFont("Helvetica-Bold", 10)
            for txt, x in [('Book Title', 50), ('Borrower', 200), ('Borrowed', 300),
                           ('Due Date', 380), ('Status', 460)]:
                p.drawString(x, y, txt)
            y -= 20
            p.setFont("Helvetica", 9)
            for item in data:
                if y < 50:
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica", 9)
                title = (item.book.title[:28] + '...') if item.book and len(item.book.title) > 30 else (item.book.title if item.book else 'N/A')
                p.drawString(50, y, title)
                p.drawString(200, y, (item.borrower.full_name[:18] if item.borrower else 'N/A'))
                p.drawString(300, y, item.borrowed_date.strftime('%Y-%m-%d') if item.borrowed_date else '')
                p.drawString(380, y, item.due_date.strftime('%Y-%m-%d') if item.due_date else '')
                p.drawString(460, y, item.status)
                y -= 15
        elif isinstance(data[0], Book):
            p.setFont("Helvetica-Bold", 10)
            for txt, x in [('Book No.', 50), ('Title', 120), ('Author', 250),
                           ('Subject', 350), ('Avail', 450), ('Total', 500)]:
                p.drawString(x, y, txt)
            y -= 20
            p.setFont("Helvetica", 9)
            for item in data:
                if y < 50:
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica", 9)
                p.drawString(50, y, item.book_number)
                p.drawString(120, y, (item.title[:18] + '...') if len(item.title) > 20 else item.title)
                p.drawString(250, y, (item.author[:13] + '...') if len(item.author) > 15 else item.author)
                p.drawString(350, y, item.subject[:10])
                p.drawString(450, y, str(item.copies_available))
                p.drawString(500, y, str(item.total_copies))
                y -= 15
        p.save()
        buf.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'{report_name}_report_{ts}.pdf')

    # ── PDF receipt generation ─────────────────────────────────────────────────

    def generate_borrow_receipt_pdf(borrowing, book, borrower, is_bulk=False, books_list=None):
        buf = io.BytesIO()
        p = canvas.Canvas(buf, pagesize=letter)
        _, height = letter
        receipt_number = f"BR{datetime.now().strftime('%Y%m%d%H%M%S')}"
        p.setFont("Helvetica-Bold", 20)
        p.drawString(50, height - 50, "ROPHINE FIELD SCHOOL")
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 80, "Library Book Borrowing Receipt")
        p.setFont("Helvetica", 10)
        p.drawString(400, height - 50, f"Receipt No: {receipt_number}")
        p.drawString(400, height - 65, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        p.line(50, height - 95, 550, height - 95)
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, height - 115, "Borrower Information:")
        p.setFont("Helvetica", 11)
        p.drawString(70, height - 135, f"Name: {borrower.full_name}")
        p.drawString(70, height - 150, f"Email: {borrower.email}")
        p.drawString(70, height - 165, f"Role: {borrower.role}")
        p.drawString(300, height - 135, f"Librarian: {current_user.full_name}")
        p.drawString(300, height - 150, f"Date Borrowed: {borrowing.borrowed_date.strftime('%Y-%m-%d')}")
        p.drawString(300, height - 165, f"Due Date: {borrowing.due_date.strftime('%Y-%m-%d')}")
        y = height - 200
        if is_bulk and books_list:
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, f"Books Borrowed ({len(books_list)} items):")
            y -= 20
            p.setFont("Helvetica-Bold", 10)
            for txt, x in [('#', 50), ('Book Number', 80), ('Title', 200),
                           ('Author', 400), ('Subject', 500)]:
                p.drawString(x, y, txt)
            y -= 15
            p.line(50, y, 550, y)
            y -= 10
            p.setFont("Helvetica", 9)
            for idx, bi in enumerate(books_list, 1):
                if y < 50:
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica", 9)
                title = (bi.title[:28] + '...') if len(bi.title) > 30 else bi.title
                author = (bi.author[:18] + '...') if len(bi.author) > 20 else bi.author
                p.drawString(50, y, str(idx))
                p.drawString(80, y, bi.book_number)
                p.drawString(200, y, title)
                p.drawString(400, y, author)
                p.drawString(500, y, bi.subject[:10])
                y -= 15
        else:
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Book Information:")
            y -= 20
            p.setFont("Helvetica", 11)
            for line in [f"Book Number: {book.book_number}", f"Title: {book.title}",
                         f"Author: {book.author}", f"Subject: {book.subject}",
                         f"Grade Level: {book.grade_level or 'N/A'}",
                         f"ISBN: {book.isbn or 'N/A'}", f"Publisher: {book.publisher or 'N/A'}"]:
                p.drawString(70, y, line)
                y -= 15
        y = max(y, 100)
        p.line(50, y - 10, 550, y - 10)
        y -= 25
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Terms and Conditions:")
        y -= 15
        p.setFont("Helvetica", 9)
        for term in ["• Books must be returned by the due date to avoid fines.",
                     "• Late returns will incur a fine of $0.50 per day per book.",
                     "• Lost or damaged books must be replaced or paid for.",
                     "• This receipt should be kept as proof of borrowing."]:
            p.drawString(70, y, term)
            y -= 12
        p.setFont("Helvetica-Oblique", 8)
        p.drawString(50, 30, "This is a computer-generated receipt. No signature required.")
        p.drawString(400, 30, "Generated by School Library System")
        p.save()
        buf.seek(0)
        return buf, receipt_number

    def generate_return_receipt_pdf(borrowing, book, borrower, is_bulk=False,
                                    returned_books_list=None, total_fine=0, fines_paid=0):
        buf = io.BytesIO()
        p = canvas.Canvas(buf, pagesize=letter)
        _, height = letter
        receipt_number = f"RR{datetime.now().strftime('%Y%m%d%H%M%S')}"
        p.setFont("Helvetica-Bold", 20)
        p.drawString(50, height - 50, "ROPHINE FIELD SCHOOL")
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 80, "Library Book Return Receipt")
        p.setFont("Helvetica", 10)
        p.drawString(400, height - 50, f"Receipt No: {receipt_number}")
        p.drawString(400, height - 65, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        p.line(50, height - 95, 550, height - 95)
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, height - 115, "Borrower Information:")
        p.setFont("Helvetica", 11)
        p.drawString(70, height - 135, f"Name: {borrower.full_name}")
        p.drawString(70, height - 150, f"Email: {borrower.email}")
        p.drawString(300, height - 135, f"Processed by: {current_user.full_name}")
        p.drawString(300, height - 150, f"Return Date: {datetime.now().strftime('%Y-%m-%d')}")
        y = height - 185
        if is_bulk and returned_books_list:
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, f"Books Returned ({len(returned_books_list)} items):")
            y -= 20
            p.setFont("Helvetica-Bold", 10)
            for txt, x in [('#', 50), ('Book Number', 80), ('Title', 200),
                           ('Due Date', 380), ('Fine', 480)]:
                p.drawString(x, y, txt)
            y -= 15
            p.line(50, y, 550, y)
            y -= 10
            p.setFont("Helvetica", 9)
            for idx, item in enumerate(returned_books_list, 1):
                if y < 50:
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica", 9)
                title = (item['book'].title[:23] + '...') if len(item['book'].title) > 25 else item['book'].title
                p.drawString(50, y, str(idx))
                p.drawString(80, y, item['book'].book_number)
                p.drawString(200, y, title)
                p.drawString(380, y, item['borrowing'].due_date.strftime('%Y-%m-%d'))
                p.drawString(480, y, f"${item['fine']:.2f}" if item['fine'] > 0 else "$0.00")
                if item['fine_paid']:
                    p.setFillColorRGB(0, 0.5, 0)
                    p.drawString(530, y, "PAID")
                    p.setFillColorRGB(0, 0, 0)
                y -= 15
            y -= 10
            p.setFont("Helvetica-Bold", 11)
            p.drawString(350, y, "Fine Summary:")
            y -= 15
            p.setFont("Helvetica", 10)
            p.drawString(370, y, f"Total Fines: ${total_fine:.2f}")
            y -= 15
            p.drawString(370, y, f"Paid Today: ${fines_paid:.2f}")
            y -= 15
            p.setFont("Helvetica-Bold", 10)
            p.drawString(370, y, f"Balance Due: ${(total_fine - fines_paid):.2f}")
        else:
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Returned Book Information:")
            y -= 20
            p.setFont("Helvetica", 11)
            for line in [f"Book Number: {book.book_number}", f"Title: {book.title}",
                         f"Author: {book.author}",
                         f"Borrowed: {borrowing.borrowed_date.strftime('%Y-%m-%d')}",
                         f"Due Date: {borrowing.due_date.strftime('%Y-%m-%d')}",
                         f"Returned: {borrowing.returned_date.strftime('%Y-%m-%d')}"]:
                p.drawString(70, y, line)
                y -= 15
            y -= 5
            p.setFont("Helvetica-Bold", 11)
            p.drawString(70, y, "Fine Information:")
            y -= 15
            p.setFont("Helvetica", 10)
            if borrowing.fine_amount > 0:
                p.drawString(90, y, f"Fine Amount: ${borrowing.fine_amount:.2f}")
                y -= 15
                p.drawString(90, y, f"Fine Paid: {'Yes' if borrowing.fine_paid else 'No'}")
            else:
                p.drawString(90, y, "No fines incurred. Thank you for returning on time!")
        p.setFont("Helvetica-Oblique", 8)
        p.drawString(50, 30, "This is a computer-generated receipt. No signature required.")
        p.drawString(400, 30, "Generated by School Library System")
        p.save()
        buf.seek(0)
        return buf, receipt_number


# ── Entry point ────────────────────────────────────────────────────────────────

app, mail = create_app()

if __name__ == '__main__':
    hostname = socket.gethostname()
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        local_ip = "Unable to get IP"

    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'

    print("\n" + "="*60)
    print("📚 SCHOOL LIBRARY MANAGEMENT SYSTEM")
    print("="*60)
    print(f"\n✅ Server running  |  debug={debug}")
    print(f"\n📱 ACCESS:")
    print(f"   • This computer : http://127.0.0.1:5000")
    print(f"   • Other devices : http://{local_ip}:5000")
    print(f"\n💻 Host: {hostname}  |  IP: {local_ip}")
    print(f"\n📋 Default login  — user: admin  |  pass: admin123")
    print(f"\n⚠️  Change the default password immediately after first login!")
    print("="*60 + "\n")

    app.run(debug=debug, host='0.0.0.0', port=5000)
